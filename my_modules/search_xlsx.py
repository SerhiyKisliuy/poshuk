from openpyxl import load_workbook
import os

class SearchInXLSX:  # Клас для пошуку в декількох файлах ексел
    def __init__(self):
        self.__maxLenRow = 0  # Максимальна довжина строки
        self.__fileNames = []  # Масив шляхів до файлів ексел
        self.__requestSearch = ""  # Пощуковий запит
        self.__resultAll = []  # Загальний результат
        self.__onFile = True  #Пошук лише в одному файлі
        self.__header = []
    def setOnFile(self, onFile: bool):
        self.__onFile = onFile

    def setFileNames(self, fileNames):  # Отримує масив щляхів до файлів ексел
        self.__fileNames = fileNames
        #print("setFileNames")
        #print(self.__fileNames)

    def setRequestSearch(self, requestSearch):  # Отримує пошуковий запит
        self.__requestSearch = requestSearch

    def __strresult(self, m):
        res = []
        for i in m:
            x = i.value  # Значення в ячейці
            if x:
                res.append(str(i.value))
            else:
                res.append('-')  # Якщо немає значення в ячейці то створюємо відступ
        return res

    def __searchInAllBooks(self):
        result = []
        lengF = len(self.__fileNames)  # Кількість книг
        for i in range(lengF):  #
            fileName = self.__fileNames[i]  # Шлях до файлу ексел
            #print("__searchInAllBook")
            #print(fileName)
            self.__searchInOneBook(fileName)  # Передаємо книгу та шлях до файлу

        #print("Result for all book")
        #print(result)
        #result = self.__resultAll
        #return result

    def __searchInOneBook(self, fileName):  #Пошук в одній книзі

        result = []  #Массив для додавання результатів
        work_book = load_workbook(fileName, data_only=True)  # завантаження книги
        lengS = len(work_book.sheetnames)  #Кількість листів в книзі
        for i in range(lengS):  #
            work_book.active = i
            sheetName = work_book.sheetnames[i]  #Отримуємо ім'я листа
            sheetBook = work_book.active  #Отримуємо активний лист (в теорії)
            resultSheet = self.__searchInOneSheet(sheetBook, sheetName, fileName)  #
            result.append(resultSheet)
        #print("Result for one book")
        #print(result)
        return result

    def __searchInOneSheet(self, sheet, sheetName, fileName):
        result = []  # Масив рядків які відповідають запиту
        filename = self.__nameFileSplit(fileName)  #Отримуэмо ымёя файлу замысть шляху
        for i in sheet:  # Перебір рядків в одному листі
            b = False  # Запис рядка в масив заборонено
            lenRow = len(i)  #Кількість ячеєк в рядку
            if lenRow > self.__maxLenRow:  #Кількісне порівняння
                self.__maxLenRow = lenRow  #Запис максимального значення
                #print("__searchInOneSheet __maxLenRow = ")
                #print(self.__maxLenRow)
            for x in i:  # Перебір ячеєк в одному рядку
                if x.value:
                    if self.__requestSearch.lower() in str(
                            x.value).lower():  # Перевірка на совпадіння слів (у нижньому регістрі) в ячейці
                        # if dat: #При пустому пошуку буде пуста таблиця
                        b = True  # Запис рядка в масив дозволено


            if b:  # Перевірка дозволу на запис рядка в масив
                r = self.__strresult(i)  # Формуємо рядок

                r = [sheetName] + r  # Додаэмо назву листа
                r = [filename[0]] + r  # Додаэмо назву файла
                if r:
                    result.append(r)  # Записуємо рядок в масив
                    self.__resultAll.append(r)
        #print("Result for one sheet")
        #print(result)
        return result


    def __nameFileSplit(self, pathFile):  #Метод для отримання ім'я файлу замість шляху
        path_dir, name_file = os.path.split(pathFile)  #Шлях каталогу та повне ім'я файлу
        filename = os.path.splitext(name_file)  #Ім'я файлу та його тип
        #print(filename)
        return filename

    def getTableDate(self):
        if self.__onFile:
            if self.__fileNames:
                self.__searchInOneBook(self.__fileNames)
                #print("getTableDate OnFile")
                #print(self.__resultAll)
                #self.__getHeader()
                #print("getTableDate OnFile")
                #print(self.__header)

                return self.__resultAll
            else:
                #print("Not fileName")
                return []
        else:
            if self.__fileNames:
                self.__searchInAllBooks()
                #self.__getHeader()

                return self.__resultAll
            else:
                #print("Not filesName")
                return []
