from openpyxl import load_workbook

def strresult(m):
    res = []
    for i in m:
        x = i.value  # Значення в ячейці
        if x:
            res.append(str(i.value))
        else:
            res.append('-')  # Якщо немає значення в ячейці то створюємо відступ
    return res

def search(filename, entry):

    wb = load_workbook(filename, data_only=True)  #завантаження книги
    #print(wb.sheetnames)
    result = []
    head = []

    for shn in range(len(wb.sheetnames)):  #перебір по таблицям (ws - Листи, wb - вся книга)
        #print(shn)
        wb.active = shn
        ws = wb.active
        #print(ws[2])
        # print('____  ' + wb.sheetnames[ws] + '  ____')  # Друкуємо назву листа
        dat = entry

        le = len(ws[1])
        for s in range(le):  #розмір рядка (кількість слів)
            head.append(str(s)) #Заповнюємо шапку номерами стовбців

        # додаєм стовбців для імен файла та листа (в початок)
        head = ['Лист'] + head
        head = ['Файл'] + head

        # print(m)  # Друкуємо першу строку таблиці
        wsd = ws

        #wsd.delete_rows(0)  # Видаляємо першу строку в таблиці

        for i in wsd:  # Перебор строк
            b = False
            for x in i:  # Перебор ячеєк
                if x.value:

                    if dat.lower() in str(x.value).lower():  # Перевірка на совпадіння слів (у нижньому регістрі)
                        #if dat: #при пустому пошуку буде пуста таблиця
                            b = True

            if b:
                r = strresult(i)  # Формуємо строку
                lis = wb.sheetnames[shn]
                r = [lis] + r   #Додаэмо назву листа
                r = [filename] + r  #Додаэмо назву файла
                if r:
                    result.append(r)  # Формуємо список строк

    return head, result



